"""Streaming SQLite-to-Excel export for one Muse research session."""

import os
import sqlite3
from pathlib import Path

from openpyxl import Workbook


EXCEL_MAX_DATA_ROWS = 1_048_575
SIGNAL_TABLES = {
    'EEG': 'eeg_logs',
    'IMU': 'imu_logs',
    'PPG': 'ppg_logs',
}


def _safe_excel_value(value):
    """Prevent user-entered metadata from becoming an Excel formula."""
    if isinstance(value, str) and value.startswith(('=', '+', '-', '@')):
        return "'" + value
    return value


def _table_exists(connection, table):
    row = connection.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
        (table,),
    ).fetchone()
    return row is not None


def _sheet_title(signal, operator, part):
    base = f'{signal.lower()}_{operator}'
    suffix = f'_{part}' if part > 1 else ''
    return (base[:31 - len(suffix)] + suffix)[:31]


def export_session_to_excel(database_path, output_path, metadata):
    """Export a session database to a memory-efficient XLSX workbook."""
    database_path = Path(database_path)
    output_path = Path(output_path)
    if not database_path.is_file():
        raise FileNotFoundError(f'No existe la base de sesión: {database_path}')

    output_path.parent.mkdir(parents=True, exist_ok=True, mode=0o700)
    temporary_path = output_path.with_suffix('.tmp.xlsx')
    workbook = Workbook(write_only=True)

    metadata_sheet = workbook.create_sheet('metadata')
    metadata_sheet.append(['campo', 'valor'])
    for key, value in metadata.items():
        metadata_sheet.append([key, _safe_excel_value(value)])

    with sqlite3.connect(database_path) as connection:
        for signal, table in SIGNAL_TABLES.items():
            if not _table_exists(connection, table):
                continue
            columns = [
                row[1]
                for row in connection.execute(f'PRAGMA table_info({table})')
                if row[1] != 'id'
            ]
            operators = [
                row[0]
                for row in connection.execute(
                    f'SELECT DISTINCT operador FROM {table} ORDER BY operador'
                )
            ]
            column_sql = ', '.join(f'"{column}"' for column in columns)
            for operator in operators:
                cursor = connection.execute(
                    f'SELECT {column_sql} FROM {table} '
                    'WHERE operador=? ORDER BY timestamp',
                    (operator,),
                )
                part = 1
                rows_in_sheet = 0
                sheet = workbook.create_sheet(
                    _sheet_title(signal, operator, part)
                )
                sheet.append(columns)
                for row in cursor:
                    if rows_in_sheet >= EXCEL_MAX_DATA_ROWS:
                        part += 1
                        rows_in_sheet = 0
                        sheet = workbook.create_sheet(
                            _sheet_title(signal, operator, part)
                        )
                        sheet.append(columns)
                    sheet.append([_safe_excel_value(value) for value in row])
                    rows_in_sheet += 1

    workbook.save(temporary_path)
    os.chmod(temporary_path, 0o600)
    os.replace(temporary_path, output_path)
    return output_path
