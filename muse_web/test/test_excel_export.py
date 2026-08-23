"""Tests for the session Excel exporter."""

import sqlite3

from openpyxl import load_workbook

from muse_web.excel_export import export_session_to_excel


def test_exports_signals_and_escapes_metadata_formulas(tmp_path):
    database = tmp_path / 'raw.sqlite'
    output = tmp_path / 'export.xlsx'
    with sqlite3.connect(database) as connection:
        connection.execute('''
            CREATE TABLE eeg_logs (
                id INTEGER PRIMARY KEY,
                operador TEXT,
                timestamp REAL,
                channel_1 REAL,
                channel_2 REAL,
                channel_3 REAL,
                channel_4 REAL,
                channel_5 REAL
            )
        ''')
        connection.execute(
            'INSERT INTO eeg_logs VALUES (1, ?, ?, ?, ?, ?, ?, ?)',
            ('operador_a', 100.5, 1.0, 2.0, 3.0, 4.0, 0.0),
        )
        connection.execute('''
            CREATE TABLE ppg_logs (
                id INTEGER PRIMARY KEY,
                operador TEXT,
                timestamp REAL,
                channel_1 REAL
            )
        ''')
        connection.execute(
            'INSERT INTO ppg_logs VALUES (1, ?, ?, ?)',
            ('operador_a', 100.5, 42.0),
        )

    export_session_to_excel(
        database,
        output,
        {'subject_code': '=DANGEROUS', 'experiment': 'Stroop'},
    )

    workbook = load_workbook(output, read_only=True, data_only=False)
    assert {'metadata', 'eeg_operador_a', 'ppg_operador_a'} <= set(
        workbook.sheetnames
    )
    metadata_rows = list(workbook['metadata'].iter_rows(values_only=True))
    eeg_rows = list(workbook['eeg_operador_a'].iter_rows(values_only=True))
    assert metadata_rows[1] == ('subject_code', "'=DANGEROUS")
    assert eeg_rows[1][0:3] == ('operador_a', 100.5, 1.0)


def test_missing_database_is_rejected(tmp_path):
    try:
        export_session_to_excel(
            tmp_path / 'missing.sqlite',
            tmp_path / 'output.xlsx',
            {},
        )
    except FileNotFoundError:
        pass
    else:
        raise AssertionError('A missing session database must be rejected')
